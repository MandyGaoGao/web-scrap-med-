# -*- coding: utf-8 -*-
"""
Created on Wed Jul 11 15:22:54 2018

@author: gaoyu
"""

from Bio import Medline
from Bio import Entrez
from openpyxl import Workbook
import openpyxl
import datetime
#wb = openpyxl.load_workbook('existing_data_file.xlsx')



def getIDlist(query):
    Entrez.email = 'gaoyunyimandy@gmail.com'
    handle = Entrez.esearch(db='pubmed', 
                            sort='relevance', 
                            retmax='3', #maximum size 10 000
                            retmode='xml', 
                            term=query)
    results = Entrez.read(handle)
    id_list = results['IdList']
    return id_list

def fetch_details(id_list):
    ids = ','.join(id_list)
    Entrez.email = 'gaoyunyimandy@gmail.com'
    handle = Entrez.efetch(db='pubmed',
                           rettype='medline',
                           retmode='text',
                           id=ids)
    results = Medline.parse(handle)
    return results

def newexcel(product,keyword):
    with open("PRODUCT.txt", "a") as text_file:
                text_file.write(product) 
            
    wb=Workbook()
    rowname=['keyword','Article Link','Title','Date','Autors','Abstract','DOI','date of search']
    ws1=wb.active
    ws1.append(rowname)
    
    list_test = getIDlist(product+' '+keyword)
    records= fetch_details(list_test)
    records=list(records)
    
    with open("DOI.txt", "r") as text_file:
        message = text_file.read()
        lst2=message.replace(' ',';')
        lst2=lst2.split(';')
    
    for record in records:
        doi=record.get('AID','?')[0].replace('[doi]','').replace('[pii]','')
        if doi.replace(' ','') not in lst2:
            
            authors=';'.join(record.get('FAU','?'))
        
            lst=[]
            lst.append(keyword)
            lst.append('https://www.ncbi.nlm.nih.gov/pubmed/'+record.get('PMID','?'))
            lst.append(record.get("TI", "?"))
            lst.append(record.get("DP","?"))
            lst.append(authors)
            lst.append(record.get("AB", "?"))
            
            lst.append(doi)
            lst.append(datetime.date.today())
            
            with open("DOI.txt", "a") as text_file:
                text_file.write(doi) 
            ws1.append(lst)
        else:
            print('repeat')
    
    new_filename = '{}.xlsx'.format(product)
    wb.save(filename = new_filename)
    
def appendexcel(product,keyword):
    wb = openpyxl.load_workbook('{}.xlsx'.format(product))
    ws1=wb.active
    
    list_test = getIDlist(product+' '+keyword)
    records= fetch_details(list_test)
    records=list(records)
    
    with open("DOI.txt", "r") as text_file:
        message = text_file.read()
        lst2=message.replace(' ',';')
        lst2=lst2.split(';')
    
    for record in records:
        doi=record.get('AID','?')[0].replace('[doi]','').replace('[pii]','')
        if doi.replace(' ','') not in lst2:
            
            authors=';'.join(record.get('FAU','?'))
        
            lst=[]
            lst.append(keyword)
            lst.append('https://www.ncbi.nlm.nih.gov/pubmed/'+record.get('PMID','?'))
            lst.append(record.get("TI", "?"))
            lst.append(record.get("DP","?"))
            lst.append(authors)
            lst.append(record.get("AB", "?"))
            
            lst.append(doi)
            lst.append(datetime.date.today())
            
            with open("DOI.txt", "a") as text_file:
                text_file.write(doi) 
            ws1.append(lst)
        else:
            print('repeat')
    wb.save('{}.xlsx'.format(product))
 
def Main():
#    number_of_search=input('key in no. of articles for each search here:')
    product=input('key in your product for search here:')
    keyword=input('key in your keyword for search here:')
#if product serached or not:
    with open("PRODUCT.txt", "r") as text_file:
        message = text_file.read()
        searched_lst=message.replace(' ',';')
        searched_lst=searched_lst.split(';')
    if product not in searched_lst:
        newexcel(product,keyword)
    else:
        appendexcel(product,keyword)
    
if __name__=='__main__':
    Main()