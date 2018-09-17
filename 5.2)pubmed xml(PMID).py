# -*- coding: utf-8 -*-
"""
Created on Wed Jul 11 15:22:54 2018

@author: gaoyu
"""

from Bio import Medline
from Bio import Entrez
from openpyxl import Workbook
import openpyxl
import datetime
import time
import itertools
start_time = time.time()
#wb = openpyxl.load_workbook('existing_data_file.xlsx')
#create PMID.txt and PRODUCT.txt
try:
   f = open('searched_PRODUCT.txt','r')
   f.close()
   
except:    
    file = open("searched_PRODUCT.txt", "w")
    file.close() 

def getIDlist(product,keyword,numberofreq):
    query=product+' '+keyword
    Entrez.email = 'gaoyunyimandy@gmail.com'
    handle = Entrez.esearch(db='pubmed', 
                            sort='relevance', 
                            retmax=numberofreq, #maximum size 10 000
                            retmode='xml', 
                            term=query)
    results = Entrez.read(handle)
    id_list = results['IdList']
    
    try:
        f = open('{}_PMID.txt'.format(product),'r')
        f.close()
   
    except:    
        file = open("{}_PMID.txt".format(product), "w")
        file.close() 
    
    with open("{}_PMID.txt".format(product), "r") as text_file:
        message = text_file.read()
        lst2=message.split(';')
    
    newid=[]
    for i in id_list:
        if i not in lst2:
            newid.append(i)
            with open("PMID.txt", "a") as text_file:
                text_file.write(i+';') 
    return newid

def fetch_details(id_list):
    ids = ','.join(id_list)
    Entrez.email = 'gaoyunyimandy@gmail.com'
    handle = Entrez.efetch(db='pubmed',
                           rettype='medline',
                           retmode='text',
                           id=ids)
    results = Medline.parse(handle)
    return results

def newexcel(product,keyword,numberofreq):
    with open("searched_PRODUCT.txt", "a") as text_file:
                text_file.write(product+',') 
            
    wb=Workbook()
    rowname=['PMID','keyword','Article Link','Title','Date','Autors','Abstract','DOI','date of search']
    ws1=wb.active
    ws1.append(rowname)
    
    list_test = getIDlist(product,keyword,numberofreq)
    records= fetch_details(list_test)
    records=list(records)

    for record in records:
        doi=record.get('AID','no article found for such keyword')[0].replace('[doi]','').replace('[pii]','')
        authors=';'.join(record.get('FAU','?'))
        lst=[]
        lst.append(record.get('PMID','?'))
        lst.append(keyword)
        lst.append('https://www.ncbi.nlm.nih.gov/pubmed/'+record.get('PMID','?'))
        lst.append(record.get("TI", "?"))
        lst.append(record.get("DP","?"))
        lst.append(authors)
        lst.append(record.get("AB", "?"))
        lst.append(doi)
        lst.append(datetime.date.today())
        ws1.append(lst)

    new_filename = '{}.xlsx'.format(product)
    wb.save(filename = new_filename)
    
def appendexcel(product,keyword,numberofreq):
    #create PMID.txt and PRODUCT.txt
    
    wb = openpyxl.load_workbook('{}.xlsx'.format(product))
    ws1=wb.active
    
    list_test = getIDlist(product,keyword,numberofreq)
    records= fetch_details(list_test)
    records=list(records)

    for record in records:
        doi=record.get('AID','?')[0].replace('[doi]','').replace('[pii]','')
        authors=';'.join(record.get('FAU','?'))
        lst=[]
        lst.append(record.get('PMID','?'))
        lst.append(keyword)
        lst.append('https://www.ncbi.nlm.nih.gov/pubmed/'+record.get('PMID','?'))
        lst.append(record.get("TI", "?"))
        lst.append(record.get("DP","?"))
        lst.append(authors)
        lst.append(record.get("AB", "?"))       
        lst.append(doi)
        lst.append(datetime.date.today())       
        ws1.append(lst)

    wb.save('{}.xlsx'.format(product))
 
def Main():
#get product list
    with open("Search_requirement.txt", "r") as text_file:
        message = text_file.read()
        requirement_lst=message.split(';')
        listofreq=[]
        for i in requirement_lst:
            listofreq.append(i.split(','))
            
    numberofreq=listofreq[0][0]
    stuff=listofreq[1]
    newproduct_lst=listofreq[2]       
    
    lst_of_keys=[]
    for L in range(1, len(stuff)+1):
        for subset in itertools.combinations(stuff, L):
            keyword=''
            for i in subset:
                
                keyword=keyword+' '+i
            
            lst_of_keys.append(keyword)
#run for excel   
    for j in lst_of_keys:
        for i in newproduct_lst:
        
            with open("searched_PRODUCT.txt", "r") as text_file:
                message = text_file.read()
                searched_lst=message.split(',')
            if i not in searched_lst:
                newexcel(i,j,numberofreq)
            else:
                appendexcel(i,j,numberofreq)
    
if __name__=='__main__':
    Main()

print("--- %s seconds ---" % (time.time() - start_time))
    
#check doi repetition:
'''   
    with open("DOI.txt", "r") as text_file:
        message = text_file.read()
        lst2=message.replace(' ',';')
        lst2=lst2.split(';')
        if doi.replace(' ','') not in lst2:
            with open("DOI.txt", "a") as text_file:
                text_file.write(doi) 
'''    